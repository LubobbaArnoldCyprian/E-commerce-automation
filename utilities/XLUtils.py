import openpyxl as O
from utilities.read_properties import ReadConfig

excel_file = "C:\\Users\\lubob\\OneDrive\\Documents\\Book1.xlsx"
excel_worksheet = "Sheet1"
wb = O.load_workbook(excel_file)
ws = wb[excel_worksheet]
row_num = ws.max_row
col_num = ws.max_column
print("the row no. is", row_num, "and the column is", col_num)
row = 2
print("product = ", ws.cell(row, 1).value)





























#
















import openpyxl
#
#     def getRowCount(file,sheetName):
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.get_sheet_by_name(sheetname)
#         return (sheet.max_row)
#
#     def getColumnCount(file, sheetName):
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.get_sheet_by_name(sheetname)
#         return (sheet.max_column)
#
#
#     def readData(file, sheetName,rownnum,columno):
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.get_sheet_by_name(sheetname)
#         return sheet.cell(row=rownum, column=columno).value
#
#
#     def writeData(file, sheetName, rownum,columno,data):
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.get_sheet_by_name(sheetname)
#         sheet.cell(row=rownum, column=columno).value = data
#         workbook.save(file)
#
#
